# -*- coding: utf-8 -*-
"""
ソルバーが出力した結果JSONファイルを読み込み、
元のExcelテンプレートにシフト情報を書き込んで、新しいExcelファイルとして出力します。

このツールはAIエージェントによる利用を想定しており、堅牢なエラーハンドリングと
自己修正を促すための豊富なエラー情報をJSON形式で標準エラー出力に提供する。

Usage:
    python result_writer.py -t <template.xlsx> -r <result.json> -m <map.json> [-o <output.xlsx>]

Args:
    -t, --template-file (str): 書式の元となるExcelファイルのパス。
    -r, --result-file   (str): shift_solverが出力した結果JSONファイルのパス。
    -m, --map-file      (str): 氏名をキー、staff_idを値とするJSONマッピングファイルのパス。
    -o, --output-file   (str, optional): 最終的な成果物となる新しいExcelファイルのパス。
                               指定されない場合、元のファイル名から自動生成されます。

Returns:
    (stdout): 成功した場合、生成されたExcelファイルのパスを含むJSONオブジェクト。
              例: {"status": "success", "output_file": "/path/to/output.xlsx"}
    (stderr): エラーが発生した場合、エラーコードとメッセージを含むJSONオブジェクト。
              例: {"status": "error", "error_code": "FILE_NOT_FOUND", ...}
"""
import utf8_utils

import sys
import os
import json
import logging
import argparse
from datetime import datetime
from dataclasses import dataclass, field
from typing import Any, Dict, List, Tuple

# --- 依存ライブラリの確認 ---
try:
    import openpyxl
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
except ImportError as e:
    print(json.dumps({
        "status": "error",
        "error_code": "DEPENDENCY_ERROR",
        "message": "必要なライブラリ'openpyxl'が見つかりません。'pip install openpyxl'を実行してください。",
        "remediation_suggestion": "Python環境にopenpyxlをインストールしてください。"
    }, ensure_ascii=False), file=sys.stderr)
    sys.exit(1)

# --- データクラス定義 ---
@dataclass
class LocationInfo:
    """Excel上の位置情報（例: スケジュール行番号）を格納するデータクラス。"""
    schedule_row: int

@dataclass
class StaffIdMapEntry:
    """name_to_id_mapの各エントリ（スタッフ名に対応する情報）を格納するデータクラス。"""
    id: str
    location: LocationInfo

@dataclass
class ShiftAssignment:
    """result.json内の各シフト割り当てを格納するデータクラス。"""
    staff_id: str
    staff_name: str
    shift_internal_name: str
    shift_display_name: str

@dataclass
class ErrorContext:
    """エラーハンドリング関数に渡す情報を集約するデータクラス。"""
    target_path: str | None = None
    exception: Exception | None = None
    details: dict[str, Any] = field(default_factory=dict)

# --- カスタム例外とArgumentParser ---
class ArgumentParsingError(Exception):
    """コマンドライン引数の解析中にエラーが発生したことを示すためのカスタム例外。"""

class GracefulArgumentParser(argparse.ArgumentParser):
    """デフォルトのエラー処理をオーバーライドし、カスタム例外を送出するArgumentParser。"""
    def error(self, message: str):
        raise ArgumentParsingError(message)

# --- 汎用エラー出力 ---
def eprint_error(error_obj: dict):
    """構造化されたエラーオブジェクトをJSON形式で標準エラー出力に出力する。"""
    print(json.dumps(error_obj, ensure_ascii=False), file=sys.stderr)

# --- エラーハンドリング関数群 ---
def handle_argument_parsing_error(context: ErrorContext):
    eprint_error({
        "status": "error",
        "error_code": "ARGUMENT_PARSING_ERROR",
        "message": "コマンドライン引数の解析に失敗しました。",
        "remediation_suggestion": "パラメータ指定を見直し、必須引数が揃っているか確認してください。",
        "details": {"original_error": str(context.exception)}
    })

def handle_file_not_found(context: ErrorContext):
    eprint_error({
        "status": "error",
        "error_code": "FILE_NOT_FOUND",
        "message": f"入力ファイルが見つかりません: {context.target_path}",
        "remediation_suggestion": "指定パスが正しいか、ファイルが存在するか確認してください。",
        "details": {"checked_path": context.target_path}
    })

def handle_json_decode_error(context: ErrorContext):
    eprint_error({
        "status": "error",
        "error_code": "JSON_DECODE_ERROR",
        "message": "JSONファイルの解析に失敗しました。ファイルが破損しているか、形式が正しくありません。",
        "remediation_suggestion": "エディタで開き、JSON構造が正しいか検証してください。",
        "details": {"error": str(context.exception)}
    })

def handle_excel_structure_error(context: ErrorContext):
    eprint_error({
        "status": "error",
        "error_code": "EXCEL_STRUCTURE_ERROR",
        "message": "Excelテンプレートの構造解析に失敗しました。予期せぬ形式の可能性があります。",
        "remediation_suggestion": "ヘッダー行や日付列が正しく配置されているか確認してください。",
        "details": {"error": str(context.exception)}
    })

def handle_unexpected_error(context: ErrorContext):
    eprint_error({
        "status": "error",
        "error_code": "UNEXPECTED_ERROR",
        "message": "ファイルの処理中に予期せぬエラーが発生しました。",
        "remediation_suggestion": "処理対象ファイルや依存環境を確認し、再実行してください。",
        "details": {"error": str(context.exception)}
    })

# --- ヘルパー関数 ---
def normalize_cell_value(value: Any) -> str:
    """
    セルの値を正規化された文字列に変換する。
    
    Args:
        value: セルの値（任意の型）
        
    Returns:
        str: 正規化された文字列（空白文字を削除済み）
    """
    if value is None:
        return ""
    return str(value).strip().replace(" ", "").replace("　", "")

# --- メインクラス ---
class ResultWriter:
    """ソルバーの結果をExcelに書き込む処理をカプセル化するクラス。"""

    def __init__(self, template_path: str, result_path: str, map_path: str):
        """
        ResultWriterインスタンスを初期化する。
        
        Args:
            template_path (str): 書式の元となるExcelファイルのパス
            result_path (str): shift_solverが出力した結果JSONファイルのパス  
            map_path (str): 氏名をキー、staff_idを値とするJSONマッピングファイルのパス
        """
        self.template_path = template_path
        self.workbook: Workbook = self._load_workbook(template_path)
        self.sheet: Worksheet = self.workbook.active
        
        # データを一括で読み込んで複数の属性に展開
        data = self._load_data(result_path, map_path)
        self.results = data[0]
        self.id_to_name_map = data[1]
        self.id_to_location_map = data[2]
        self.original_input_filename = data[3]
        self.day_col_map = data[4]

    def _load_workbook(self, path: str) -> Workbook:
        """
        指定されたパスからExcelワークブックを読み込む。
        
        Args:
            path (str): Excelファイルのパス
            
        Returns:
            Workbook: 読み込まれたワークブック
        """
        logging.info("テンプレートExcel '%s' を読み込んでいます...", path)
        return openpyxl.load_workbook(path)

    def _parse_shift_assignments(self, raw_results: Dict[str, Any]) -> Dict[str, List[ShiftAssignment]]:
        """
        生の結果データからShiftAssignmentオブジェクトのリストを生成する。
        
        Args:
            raw_results: 生の結果データ辞書
            
        Returns:
            Dict[str, List[ShiftAssignment]]: 日付別のシフト割り当て辞書
        """
        parsed_results: Dict[str, List[ShiftAssignment]] = {}
        for date_str, assignments_data in raw_results.items():
            parsed_assignments: List[ShiftAssignment] = []
            for assign_data in assignments_data:
                try:
                    parsed_assignments.append(ShiftAssignment(**assign_data))
                except TypeError as e:
                    logging.warning("シフト割り当てデータが不正な形式です: %s. スキップします。エラー: %s", assign_data, e)
            parsed_results[date_str] = parsed_assignments
        return parsed_results

    def _parse_staff_mappings(self, raw_name_to_id_map: Dict[str, Any]) -> Tuple[Dict[str, str], Dict[str, LocationInfo]]:
        """
        生のマッピングデータからスタッフのID→氏名と位置情報のマッピングを生成する。
        
        Args:
            raw_name_to_id_map: 生のマッピングデータ辞書
            
        Returns:
            Tuple[Dict[str, str], Dict[str, LocationInfo]]: ID→氏名とID→位置情報のマッピング
        """
        parsed_name_to_id_map: Dict[str, StaffIdMapEntry] = {}
        for name, data in raw_name_to_id_map.items():
            if name.startswith("_"):  # メタデータセクションはスキップ
                continue
            if not isinstance(data, dict) or 'id' not in data or 'location' not in data:
                logging.warning("マッピングエントリ '%s' が不正な形式です。スキップします。", name)
                continue
            location_data = data['location']
            if not isinstance(location_data, dict) or 'schedule_row' not in location_data:
                logging.warning("マッピングエントリ '%s' の位置情報が不正な形式です。スキップします。", name)
                continue
            try:
                location_info = LocationInfo(**location_data)
                parsed_name_to_id_map[name] = StaffIdMapEntry(id=data['id'], location=location_info)
            except TypeError as e:
                logging.warning("マッピングエントリ '%s' のデータが不正な形式です。スキップします。エラー: %s", name, e)

        id_to_name_map = {entry.id: name for name, entry in parsed_name_to_id_map.items()}
        id_to_location_map = {entry.id: entry.location for name, entry in parsed_name_to_id_map.items()}
        return id_to_name_map, id_to_location_map

    def _extract_metadata(self, raw_name_to_id_map: Dict[str, Any]) -> Tuple[str, Dict[str, int]]:
        """
        生のマッピングデータからメタデータ（ファイル名とレイアウトヒント）を抽出する。
        
        Args:
            raw_name_to_id_map: 生のマッピングデータ辞書
            
        Returns:
            Tuple[str, Dict[str, int]]: 元のファイル名と日付→列インデックスマッピング
        """
        # _file_info から original_input_filename を取得
        file_info = raw_name_to_id_map.get("_file_info", {})
        original_input_filename = file_info.get("original_input_filename")
        if not original_input_filename:
            raise ValueError("マッピングファイルに元のファイル名情報（_file_info.original_input_filename）が見つかりません。")

        # _excel_layout_hints から day_column_map を取得
        layout_hints = raw_name_to_id_map.get("_excel_layout_hints", {})
        day_column_map = layout_hints.get("day_column_map", {})
        if not day_column_map:
            raise ValueError("マッピングファイルにExcelのレイアウトヒント（day_column_map）が見つかりません。")
            
        return original_input_filename, day_column_map
    def _load_data(self, result_path: str, map_path: str) -> Tuple[Dict[str, List[ShiftAssignment]], Dict[str, str], Dict[str, LocationInfo], str, Dict[str, int]]:
        """
        結果JSONファイルとマッピングファイルからデータを読み込む。
        
        Args:
            result_path (str): 結果JSONファイルのパス
            map_path (str): マッピングJSONファイルのパス
            
        Returns:
            Tuple containing:
                - Dict[str, List[ShiftAssignment]]: 日付別のシフト割り当て
                - Dict[str, str]: ID→氏名マッピング
                - Dict[str, LocationInfo]: ID→位置情報マッピング
                - str: 元の入力ファイル名
                - Dict[str, int]: 日付→列インデックスマッピング
        """
        logging.info("データファイル '%s' と '%s' を読み込んでいます...", result_path, map_path)
        
        with open(result_path, 'r', encoding='utf-8') as f:
            raw_results = json.load(f)
        with open(map_path, 'r', encoding='utf-8') as f:
            raw_name_to_id_map = json.load(f)

        # 各種データを別メソッドで解析
        parsed_results = self._parse_shift_assignments(raw_results)
        id_to_name_map, id_to_location_map = self._parse_staff_mappings(raw_name_to_id_map)
        original_input_filename, day_column_map = self._extract_metadata(raw_name_to_id_map)

        logging.info("%d 件のシフト割り当てを読み込みました。", sum(len(v) for v in parsed_results.values()))
        logging.info("%d 件のID→氏名マッピングを作成しました。", len(id_to_name_map))
        logging.info("%d 件のID→位置マッピングを作成しました。", len(id_to_location_map))
        logging.info("%d 日分の列マッピングを読み込みました。", len(day_column_map))
        return parsed_results, id_to_name_map, id_to_location_map, original_input_filename, day_column_map

    def write_shifts(self):
        """解析結果をExcelシートオブジェクトに書き込む。"""
        logging.info("シフト情報の書き込みを開始します...")
        written_count = 0
        for date_str, assignments in self.results.items():
            try:
                day = datetime.strptime(date_str, "%Y-%m-%d").day
            except ValueError:
                logging.warning("日付形式が不正です: %s。スキップします。", date_str)
                continue

            # day_col_map のキーが文字列なので、day を文字列に変換して検索
            if str(day) not in self.day_col_map:
                logging.warning("%d日の列がExcelに見つかりません。", day)
                continue
            
            col_idx = self.day_col_map[str(day)]  # day を文字列に変換してキーとして使用

            for assignment in assignments:
                staff_id = assignment.staff_id
                shift_display_name = assignment.shift_display_name

                # IDに対応する位置情報が存在するかどうかを直接確認し、警告を改善
                location_info = self.id_to_location_map.get(staff_id)
                if not location_info or not isinstance(location_info, LocationInfo):
                    logging.warning("ID '%s' に対応する位置情報が見つからないか不完全です。", staff_id)
                    continue
                
                row_idx = location_info.schedule_row
                
                # col_idx に +1 を追加 (0-based → 1-based)
                self.sheet.cell(row=row_idx, column=col_idx + 1, value=shift_display_name)
                written_count += 1
        
        logging.info("合計 %d 件のシフトをExcelに書き込みました。", written_count)

    def save(self, output_path: str):
        """
        ワークブックを指定されたパスに保存する。
        
        Args:
            output_path (str): 保存先のファイルパス
        """
        output_dir = os.path.dirname(output_path)
        if output_dir:
            os.makedirs(output_dir, exist_ok=True)
        self.workbook.save(output_path)
        logging.info("結果を '%s' に保存しました。", output_path)

# --- メイン実行部 ---
def main():
    """
    スクリプトのメインエントリーポイント。
    コマンドライン引数を解析し、結果の書き込み処理を実行する。
    """
    logging.basicConfig(level=logging.INFO, stream=sys.stderr, format='%(levelname)s: %(message)s')

    parser = GracefulArgumentParser(description="ソルバーの計算結果(JSON)を、元のExcelファイルに書き込みます。")
    parser.add_argument("-t", "--template-file", required=True, help="書式の元となるExcelファイルのパス。")
    parser.add_argument("-r", "--result-file", required=True, help="shift_solverが出力した結果JSONファイルのパス。")
    
    parser.add_argument("-m", "--map-file", required=True, help="氏名をキー、staff_idを値とするJSONマッピングファイルのパス。")

    try:
        args = parser.parse_args()
        
        writer = ResultWriter(
            template_path=args.template_file,
            result_path=args.result_file,
            map_path=args.map_file
        )
        
        # 出力ファイル名を決定
        # 元のファイル名から拡張子を除いた部分を取得し、"_完成.xlsx"を追加
        base_name = os.path.splitext(writer.original_input_filename)[0]
        output_filename = f"{base_name}_完成.xlsx"
        output_file = os.path.join("Output", output_filename)
        
        writer.write_shifts()
        writer.save(output_file)
        print(json.dumps({
            "status": "success",
            "output_file": os.path.abspath(output_file)
        }, ensure_ascii=False))
        
    except ArgumentParsingError as e:
        handle_argument_parsing_error(ErrorContext(exception=e))
        sys.exit(1)
    except FileNotFoundError as e:
        handle_file_not_found(ErrorContext(exception=e, target_path=e.filename))
        sys.exit(1)
    except json.JSONDecodeError as e:
        handle_json_decode_error(ErrorContext(exception=e))
        sys.exit(1)
    except ValueError as e:
        handle_excel_structure_error(ErrorContext(exception=e))
        sys.exit(1)
    except Exception as e:
        handle_unexpected_error(ErrorContext(exception=e))
        sys.exit(1)

if __name__ == '__main__':
    main()
